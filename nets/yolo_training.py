# CTK 修改内容：将Cosface或者Arcface融入到 YOLOLoss这个类里面 里面

# CTK 修改效果
from random import shuffle
import math
import cv2
import numpy as np
import torch
import torch.nn as nn
import torch.nn.functional as F
from matplotlib.colors import hsv_to_rgb, rgb_to_hsv
from PIL import Image
from utils.utils import bbox_iou
from utils.utils_decode_pred_pos import DecodeBox
import os
from utils.config import Config
import math
import torch.nn.functional as F
import torchvision
# from core.config import cfg
from openpyxl import Workbook
from openpyxl import load_workbook
# 防止print的过程中由于字符串的长度过长，显示有省略号的问题。
np.set_printoptions(threshold=999999999)


os.environ['DISPLAY'] = ':0'


def jaccard(_box_a, _box_b):
    # 计算真实框的左上角和右下角
    b1_x1, b1_x2 = _box_a[:, 0] - _box_a[:, 2] / 2, _box_a[:, 0] + _box_a[:, 2] / 2
    b1_y1, b1_y2 = _box_a[:, 1] - _box_a[:, 3] / 2, _box_a[:, 1] + _box_a[:, 3] / 2
    # 计算先验框的左上角和右下角
    b2_x1, b2_x2 = _box_b[:, 0] - _box_b[:, 2] / 2, _box_b[:, 0] + _box_b[:, 2] / 2
    b2_y1, b2_y2 = _box_b[:, 1] - _box_b[:, 3] / 2, _box_b[:, 1] + _box_b[:, 3] / 2
    box_a = torch.zeros_like(_box_a)
    box_b = torch.zeros_like(_box_b)
    box_a[:, 0], box_a[:, 1], box_a[:, 2], box_a[:, 3] = b1_x1, b1_y1, b1_x2, b1_y2
    box_b[:, 0], box_b[:, 1], box_b[:, 2], box_b[:, 3] = b2_x1, b2_y1, b2_x2, b2_y2
    A = box_a.size(0)
    B = box_b.size(0)
    max_xy = torch.min(box_a[:, 2:].unsqueeze(1).expand(A, B, 2),
                       box_b[:, 2:].unsqueeze(0).expand(A, B, 2))
    min_xy = torch.max(box_a[:, :2].unsqueeze(1).expand(A, B, 2),
                       box_b[:, :2].unsqueeze(0).expand(A, B, 2))
    inter = torch.clamp((max_xy - min_xy), min=0)

    inter = inter[:, :, 0] * inter[:, :, 1]
    # 计算先验框和真实框各自的面积
    area_a = ((box_a[:, 2]-box_a[:, 0]) *
              (box_a[:, 3]-box_a[:, 1])).unsqueeze(1).expand_as(inter)  # [A,B]
    area_b = ((box_b[:, 2]-box_b[:, 0]) *
              (box_b[:, 3]-box_b[:, 1])).unsqueeze(0).expand_as(inter)  # [A,B]
    # 求IOU
    union = area_a + area_b - inter
    return inter / union  # [A,B]
    
def clip_by_tensor(t,t_min,t_max):
    t=t.float()
 
    result = (t >= t_min).float() * t + (t < t_min).float() * t_min
    result = (result <= t_max).float() * result + (result > t_max).float() * t_max
    return result

def MSELoss(pred,target):
    return (pred-target)**2

def BCELoss(pred,target):
    epsilon = 1e-7
    pred = clip_by_tensor(pred, epsilon, 1.0 - epsilon)
    output = -target * torch.log(pred) - (1.0 - target) * torch.log(1.0 - pred)
    return output

class YOLOLoss(nn.Module):
    # anchor的形式为（9，2）
    # num_classes 为 5
    # img_size为（416,416）
    # Cuda是否为启用加速
    # normalize是否对loss的输出值进行归一化，也就是说是否除以batch。
    # num_in_feature也就是embedding，对于人脸识别来说，128维度的向量来代表一张人脸特征。
    # # 数据集中类别的数量（例如，URPC中为5， COCO为80）。
    def __init__(self, anchors, num_classes, img_size, cuda, normalize, num_in_features, num_out_features, center_update_coef, s, m):

        super(YOLOLoss, self).__init__()
        #-----------------------------------------------------------#
        #   CTK  在init函数将shape为【9，2】的anchors全部传入，不同特征层的anchor如下所示
        #   13x13的特征层对应的anchor是[116,90],[156,198],[373,326]
        #   26x26的特征层对应的anchor是[30,61],[62,45],[59,119]
        #   52x52的特征层对应的anchor是[10,13],[16,30],[33,23]
        #-----------------------------------------------------------#
        self.anchors = anchors # 此处anchor的shape 为【9，2】
        self.num_anchors = len(anchors) # CTK self.num_anchors的长度为9
        self.num_classes = num_classes
        self.bbox_attrs = 5 + num_classes
        #-------------------------------------#
        #   获得特征层的宽高
        #   13、26、52
        #-------------------------------------#
        self.feature_length = [img_size[0]//32,img_size[0]//16,img_size[0]//8]     # CTK 416//32, 416//16, 416//8
        self.img_size = img_size

        self.ignore_threshold = 0.5
        self.lambda_xy = 1.0
        self.lambda_wh = 1.0
        self.lambda_conf = 1.0
        self.lambda_cls = 1.0
        self.cuda = cuda
        self.normalize = normalize
        self.config = Config

        #----------------------------------#
        # 定义arcface或者cosface的初始参数
        self.num_in_features = num_in_features    # 表示 num_in_feature也就是embedding，对于人脸识别来说，128维度的向量来代表一张人脸特征。
        self.num_out_features = num_out_features  # 数据集中类别的数量（例如，URPC中为5， COCO为80）。
        self.s = s # 缩放因子
        self.m = m # 在cosface中，为cos（theta）-self.m做好准备。
        self.weight = nn.Parameter(torch.FloatTensor(self.num_out_features, self.num_in_features)) # 定义权重向量
        self.coefficient = torch.nn.Parameter(torch.Tensor([0.0]))
        nn.init.xavier_uniform_(self.weight)  # 表示 均匀分布
        #----------------------------------#

        # CTK 修改
        self.img_size_width  = img_size[0]
        self.img_size_height = img_size[1]

        # CTK 修改
        self.avg = nn.AdaptiveAvgPool2d((1, 1))
        #self.max = nn.AdaptiveMaxPool2d((1, 1))

        self.yolo_decodes = []
        for i in range(3):
            # yolo_decodes[0]的形式为（b, 1, 13*13, 4），对应最大的anchors
            # yolo_decodes[1]的形式为（b, 1, 26*26, 4），对应中等的anchors
            # yolo_decodes[2]的形式为（b, 1, 52*52, 4），对应最小的anchors
            self.yolo_decodes.append(DecodeBox(self.config["yolo"]["anchors"][i], self.config["yolo"]["classes"], (self.img_size_width, self.img_size_height)))

        # 定义变量，分别存储每一个epoch中 echinus、holothurian、starfish、scallop的特征向量
        # ['echinus', 'holothurian', 'starfish', 'scallop', 'waterweeds']
        self.global_embeddings_echinus = torch.zeros([1, self.num_in_features])
        self.global_embeddings_holothurian = torch.zeros([1, self.num_in_features])
        self.global_embeddings_starfish = torch.zeros([1, self.num_in_features])
        self.global_embeddings_scallop = torch.zeros([1, self.num_in_features])

        # 定义全局变量来保存类别中心（目的是：通过使用滑动机制，使得最新一次类别中心的更新与上次有一定的关系，防止更新的过快。）
        self.cls_echinus_global_center = torch.zeros([1, self.num_in_features])
        self.cls_holothurian_global_center = torch.zeros([1, self.num_in_features])
        self.cls_starfish_global_center = torch.zeros([1, self.num_in_features])
        self.cls_scallop_global_center = torch.zeros([1, self.num_in_features])

        # 定义每个类别的方差值
        self.var_echinus = torch.Tensor([0.0])
        self.var_holothurian = torch.Tensor([0.0])
        self.var_starfish = torch.Tensor([0.0])
        self.var_scallop = torch.Tensor([0.0])

        # 每次epoch后，每个item与中心的夹角
        self.angular_between_items_with_center_echinus = []
        self.angular_between_items_with_center_holothurian = []
        self.angular_between_items_with_center_starfish = []
        self.angular_between_items_with_center_scallop = []


        # 类别中心更新率
        self.center_update_coef = center_update_coef


        #----------------------------------------------------#
        #   CTK 确定input的shape为下述三行的  某一行
        #   bs, 3*(5+num_classes), 13, 13
        #   bs, 3*(5+num_classes), 26, 26
        #   bs, 3*(5+num_classes), 52, 52
        #----------------------------------------------------#
        #----------------------------------------------------#
        #   CTK 确定 feature_map 的shape为下述三行中的某一行，表示特征图,此处并不是 embeddings
        #   bs, 128, 13, 13
        #   bs, 128, 26, 26
        #   bs, 128, 52, 52
        #----------------------------------------------------#
        # CTK  target的总数量为bs，对于bs的中每一个batch，其形式为N*5，其中N代表本张图片里面有多少个物体，5代表x,y,w,h和类别
        # input的shape为【bs】
    def forward(self, input, feature_map, targets, epoch, is_save):
        #print('feature_map.shape', feature_map.shape)
        #-----------------------#
        #   一共多少张图片
        #-----------------------#
        bs = input.size(0)
        #-----------------------#
        #   特征层的高
        #-----------------------#
        in_h = input.size(2)
        #-----------------------#
        #   特征层的宽
        #-----------------------#
        in_w = input.size(3)

        # CTK 修改
        # 首先对input输入进行解码工作，
        # CTK 确定 self.feature_length=[13、26、52], 因此借助于index函数，最终 subtract_index 的值只能为【0，1，2】中任意一个值
        # 因此，若本次遍历的是13*13特征层，则 feature_map_index 对应【0】
        #      若本次遍历的是26*26特征层，则 feature_map_index 对应【1】
        #      若本次遍历的是52*52特征层，则 feature_map_index 对应【2】

        feature_map_index = [0,1,2][self.feature_length.index(in_w)]


        # decode_output的输出为
        #             # decode_output[0]的形式为（b, 3, 13*13, 4），对应最大的anchors
        #             # decode_output[1]的形式为（b, 3, 26*26, 4），对应中等的anchors
        #             # decode_output[2]的形式为（b, 3, 52*52, 4），对应最小的anchors
        # 等号右侧： input的shape为（bs, 3*(5+num_classes), 13, 13）
        # 等号左侧： decode_output 的 shape为 (bs，3, 13, 13, 4) 其中4代表x,y,w,h,其是相对于特征图尺寸。
        decode_output = self.yolo_decodes[feature_map_index](input)

        #-----------------------------------------------------------------------#
        #   计算步长
        #   每一个特征点对应原来的图片上多少个像素点
        #   如果特征层为13x13的话，一个特征点就对应原来的图片上的32个像素点
        #   如果特征层为26x26的话，一个特征点就对应原来的图片上的16个像素点
        #   如果特征层为52x52的话，一个特征点就对应原来的图片上的8个像素点
        #   stride_h = stride_w = 32、16、8
        #-----------------------------------------------------------------------#
        stride_h = self.img_size[1] / in_h  #  stride_h 为 32、16、8按照顺序其中的一个
        stride_w = self.img_size[0] / in_w

        #-------------------------------------------------#
        #   此时获得的scaled_anchors大小是相对于特征层的
        #-------------------------------------------------#
        # CTK 确定等号右侧 self.anchors的shape是【9，2】，原先的（a_w，a_h）是相对于416*416图片大小，
        # CTK 确定等号左侧 scaled_anchors的shape为【9，2】，# 现在要将其分别缩小（stride_w， stride_h）倍
        scaled_anchors = [(a_w / stride_w, a_h / stride_h) for a_w, a_h in self.anchors]

        
        #-----------------------------------------------#
        #  CTK 等号右侧的 input是下面三行的  "某一行"
        #   bs, 3*(5+num_classes), 13, 13
        #   bs, 3*(5+num_classes), 26, 26
        #   bs, 3*(5+num_classes), 52, 52

        #  CTK 等号左侧的 prediction 是下面三行的  "某一行"
        #   bs, 3, 13, 13, 5 + num_classes
        #   bs, 3, 26, 26, 5 + num_classes
        #   bs, 3, 52, 52, 5 + num_classes
        #-----------------------------------------------#
        prediction = input.view(bs, int(self.num_anchors/3), self.bbox_attrs, in_h, in_w).permute(0, 1, 3, 4, 2).contiguous()
        
        # 先验框的中心位置的调整参数 形状为 （bs, 3, 13, 13） 或者 （bs, 3, 26, 26） 或者 （bs, 3, 52, 52）
        x = torch.sigmoid(prediction[..., 0])
        y = torch.sigmoid(prediction[..., 1])

        # 先验框的宽高调整参数 形状为 （bs, 3, 13, 13） 或者 （bs, 3, 26, 26） 或者 （bs, 3, 52, 52）
        w = prediction[..., 2]
        h = prediction[..., 3]

        # 获得置信度，是否有物体 形状为 （bs, 3, 13, 13） 或者 （bs, 3, 26, 26） 或者 （bs, 3, 52, 52）
        conf = torch.sigmoid(prediction[..., 4])

        # 种类置信度 （bs, 3, 13, 13，num_cls） 或者 （bs, 3, 26, 26, num_cls） 或者 （bs, 3, 52, 52, num_cls）
        pred_cls = torch.sigmoid(prediction[..., 5:])

        #---------------------------------------------------------------#
        #   找到哪些先验框内部包含物体
        #   利用真实框和先验框计算交并比
        #   mask        bs, 3, in_h, in_w   无目标的特征点
        #   noobj_mask  bs, 3, in_h, in_w   有目标的特征点
        #   tx          bs, 3, in_h, in_w   中心x偏移情况
        #   ty          bs, 3, in_h, in_w   中心y偏移情况
        #   tw          bs, 3, in_h, in_w   宽高调整参数的真实值
        #   th          bs, 3, in_h, in_w   宽高调整参数的真实值
        #   tconf       bs, 3, in_h, in_w   置信度真实值
        #   tcls        bs, 3, in_h, in_w, num_classes  种类真实值
        #----------------------------------------------------------------#
        # 输入 target的总数量为bs，对于bs的中每一个batch，其形式为N*5，其中N代表本张图片里面有多少个物体，5代表x,y,w,h和类别
        # scaled_anchors 是先验框缩放stride倍（也就是将先验框缩放至当前特征图大小（13*13 或 26*26 或52*52）
        # in_w 和 in_h是当前特征图的宽度和高度,(13,13)或（26,26）或（52,52）
        mask, noobj_mask, tx, ty, tw, th, tconf, tcls, box_loss_scale_x, box_loss_scale_y =\
                                                                            self.get_target(targets, scaled_anchors,
                                                                                            in_w, in_h,
                                                                                       self.ignore_threshold)


        #------------------------------------------------------------------#
        # 计算 不同类别不均衡
        # 数据0: feature_map 的shape为 （bs, 128, 13,13）表示检测头前面的一个特征层
        # 数据1：decode_output 的 shape为 (bs，3, 13, 13, 4) 其中4代表x,y,w,h,其是相对于特征图尺寸(预测值)。
        # 数据2：mask 的shape为(bs, 3, 13, 13) 表示存在目标的特征点
        # 数据3：tcls 的shape为 （bs, 3, 13, 13, 5），其中5代表5个类别，echinus, scallop, starfish, holothurian, waterweeds，其表示真实物体的类别
        # 等号左侧：cosine_value的shape为【num_objects, 5】代表cosine值，
        # 等号左侧：cls_label的shape为 【num_objects，】代表该框框所包含的真实类别值
        # 等号左侧：container_embeddings的shape为【num_objects,256】
        cosine_value, cls_label, container_embeddings = self.metric_cosine(feature_map, decode_output, mask, tcls)


        #------------------------------------------------------------------#

        #---------------------------------------------------------------#
        #   将预测结果进行解码，判断预测结果和真实值的重合程度，如果重合程度过大则忽略，因为这些特征点属于预测比较准确的特征点，作为负样本不合适
        #----------------------------------------------------------------#
        # CTK  prediction的形式为 （b, 3, 13,13, (1+ 4+ num_class)）
        # CTK  target的总数量为bs，对于bs的中每一个batch，其形式为【N，5】，其中N代表一张图片里面包含物体的个数，5代表x,y,w,h和类别，另外 targets是一个batch的目标
        #       位置信息（每张图片中所包含物体 左上角 和 右下角 坐标信息依照原始图片大小（随便举个例子，例如，325*856，总之，这里并不是416*416）被归一化0-1之间
        # CTK  scaled_anchors的shape为【9，2】，scaled_anchors大小是相对于特征层的
        # in_w 特征层的宽
        # in_h 特征层的高
        # 等号右侧的noobj_mask 的shape为（bs, 3, 13,13）
        noobj_mask = self.get_ignore(prediction, targets, scaled_anchors, in_w, in_h, noobj_mask)

        if self.cuda:
            box_loss_scale_x = (box_loss_scale_x).cuda()
            box_loss_scale_y = (box_loss_scale_y).cuda()
            mask, noobj_mask = mask.cuda(), noobj_mask.cuda()
            tx, ty, tw, th = tx.cuda(), ty.cuda(), tw.cuda(), th.cuda()
            tconf, tcls = tconf.cuda(), tcls.cuda()

        # box_loss_scale_x和box_loss_scale_y都是被归一化（0，1）之间
        box_loss_scale = 2 - box_loss_scale_x * box_loss_scale_y

        # 计算中心偏移情况的loss，使用BCELoss效果好一些
        # CTK x, tx, box_loss_scale, mask的shape均为(bs, 3, 13, 13）
        # x 和 y 是先验框的中心位置的调整参数
        loss_x = torch.sum(BCELoss(x, tx) * box_loss_scale * mask)
        loss_y = torch.sum(BCELoss(y, ty) * box_loss_scale * mask)


        # 计算宽高调整值的loss
        # CTK w, tw, box_loss_scale, mask的shape均为(bs, 3, 13, 13）
        loss_w = torch.sum(MSELoss(w, tw) * 0.5 * box_loss_scale * mask)
        loss_h = torch.sum(MSELoss(h, th) * 0.5 * box_loss_scale * mask)

        # 计算置信度的loss
        # CTK 等号右侧 conf, mask, noobj_mask 的 shape均为(bs, 3, 13, 13）
        # 其中，等号右侧 torch.sum(BCELoss(conf, mask) * mask)表示为 “负责预测海底生物的先验框（这就是乘以mask的原因）所产生的置信度预测值与真实框置信度（1）的差距”
        # 另外，等号右侧torch.sum(BCELoss(conf, mask) * noobj_mask表示为 “负责预测海底背景的先验框（这就是乘noobj_mask的原因）”所产生的置信度预测值与海底背景置信度（0）的差距“
        loss_conf = torch.sum(BCELoss(conf, mask) * mask) + \
                    torch.sum(BCELoss(conf, mask) * noobj_mask)


        # -------------------------Bubbliiiing的计算类别损失的方法------------------
        # CTK 计算类别置信度，需要注意的是，这里仅计算mask ==1 位置处类别损失
        # 对于13*13的特征层， pred_cls的shape为（bs, 3, 13, 13，5）, tcls 的shape为 （bs, 3, 13, 13, 5）
        # mask 的 shape为（bs, 3, 13, 13）
        # 因此， pred_cls[mask == 1]可以得到shape为【BBB，5】的数据，
        # 注意到上述BBB的值与在该尺度下（13*13 或者 26*26 或者 52*52）先验框匹配到真实框的个数是一致的（因为yolov3中每个真实框只能由一个先验框负责预测），
        #                                                                 并且BBB的数值并不一定等于batch size张图片中包含所有真实框的个数。
        # 例如，batch size张图片中共包含100个真实框，则在 13*13，26*26，52*52下匹配到真实框的个数分别为：
        # 在（13*13）的尺度下，batch size张图片中一共匹配到20个真实框
        # 在（26*26）的尺度下，batch size张图片中一共匹配到35个真实框
        # 在（52*52）的尺度下，batch size张图片中一共匹配到45个真实框
        # print('pred_cls[mask == 1].device', pred_cls[mask == 1].device)
        # print('tcls[mask == 1].device', tcls[mask == 1].device)

        # 陈廷凯暂时屏蔽 bubbliiiig写的计算类别损失的函数
        loss_cls = torch.sum(BCELoss(pred_cls[mask == 1], tcls[mask == 1]))

        #print('pred_cls[mask == 1].shape 本尺度下存在真值框的个数为：', pred_cls[mask == 1].shape[0])
        # print('pred_cls[mask == 1] 本尺度物体类别的独热编码为：', tcls[mask == 1])
        # print('cls_label 余弦值的类别为', cls_label)


        # ------------------------陈廷凯自己设计类别损失方法---------------
        loss_cosine = self.cosine_loss_func(cosine_value, cls_label)
        #print('loss_cosine.device', loss_cosine.device)

        #-------------------------陈廷凯 long tailed 损失函数-----------
        loss_long_tail = self.loss_long_tail_func(cls_label, container_embeddings, epoch, in_h, is_save)




        # 将上述所有的损失（中心、宽高、置信度、类别）
        #loss = loss_x * self.lambda_xy + loss_y * self.lambda_xy + \
         #      loss_w * self.lambda_wh + loss_h * self.lambda_wh + \
         #      loss_conf * self.lambda_conf + loss_cls * self.lambda_cls

        # 将上述所有的损失（中心、宽高、置信度、类别）
        loss = loss_x * self.lambda_xy + loss_y * self.lambda_xy + \
               loss_w * self.lambda_wh + loss_h * self.lambda_wh + \
               loss_conf * self.lambda_conf + \
               loss_cls * self.lambda_cls + \
               loss_cosine + \
               loss_long_tail

        if self.normalize:
            # CTK  torch.sum()对输入的tensor数据的某一维度求和，如果不指定的dim的话，则求取全部的和,表示一共有多少个正样本?
            # 因为mask里面，既包含1（代表网格中所包含三个先验框中的某一个负责预测）也包含0 （不负责预测）
            num_pos = torch.sum(mask)
            print('num_pos1',num_pos)

            num_pos = torch.max(num_pos, torch.ones_like(num_pos))
            print('num_pos2',num_pos)
        else:
            num_pos = bs/3
        return loss, num_pos

    def calculate_angular_between_each_item_and_center(self, different_cls_embedding, global_center):
        # --------------------------开始 计算echinus的方差-----------------------------------#
        #print('different_cls_embedding', different_cls_embedding.shape)
        # 得到echinus的类别中心，center_echinus的shape是【256，】
        center_different_cls_embedding = torch.mean(different_cls_embedding, dim=0).unsqueeze(dim=0)
        update_center = center_different_cls_embedding * (1-self.center_update_coef) + global_center * self.center_update_coef

        # 计算每一行元素的平方值（包括中心行的平方值 和 每一个样本的平方值）
        # 等号右侧：center_echinus的shape是（1, 256）
        # 等号右侧：all_tensor_echinus_pow2的shape为【AAA， 256】,这个AAA我也不知道是多少，表示已经获得echinus的个数
        center_pow2 = update_center ** 2
        all_tensor_pow2 = different_cls_embedding ** 2

        # 计算分子乘积,
        # 等号右侧：center_echinus的shape是（1, 256）
        # 等号右侧：all_tensor_echinus_pow2的shape为【AAA， 256】,这个AAA我也不知道是多少，表示已经获得echinus的个数
        # 等号左侧：echinus_fenzi利用广播机制，未进行torch.sum操作之前，其shape为【AAA，256】； 利用torch.sum，得到【AAA，】
        fenzi = torch.sum(update_center * different_cls_embedding, dim=1)

        # 计算分母,
        # 等号右侧：(torch.sum(center_echinus_pow2, dim=1)) ** 0.5 计算中心特征的长度
        # 等号右侧：(torch.sum(all_tensor_echinus_pow2, dim=1)) ** 0.5，计算每一行特征的长度
        fenmu = (torch.sum(center_pow2, dim=1)) ** 0.5 * (torch.sum(all_tensor_pow2, dim=1)) ** 0.5

        # 计算echinus中每一个向量与中心向量的角度, angular_echinus的shape为[AAA，], 这个AAA我也不知道是多少，表示已经获得echinus的个数
        angular = torch.acos(fenzi / fenmu, out=None) * 180 / math.pi

        # 查看angular中是否存在nan的数值，
        # 等号左侧：pos_nan_bool的shape为[AAA，]，但是这AAA个值都是由True或者False组成的。如果该位置是nan，则返回True，否则返回Fasle
        pos_nan_bool = torch.isnan(angular)

        # 等号右侧：利用取反操作‘~’
        # 等号左侧：aft_nan_process的shape为【BBB，】，其中BBB的数量我也不知道是多少
        aft_nan_process = angular[~pos_nan_bool]

        return aft_nan_process, update_center

    # decending_var_num的shape为【4,2】，第一行：【最大的方差，类别数量】， 第二行：【第二大方差，类别数量】， 第三行：【第三大方差，类别数量】， 第四行：【第四大方差，类别数量】
    def calculate_diff_cls_loss(self, data):

        loss_first  = (torch.Tensor([0.]))
        loss_second = (data[0,0]-data[1,0])
        loss_third  = (data[0,0]-data[2,0]) + (data[1,0]-data[2,0])
        loss_fourth = (data[0,0]-data[3,0]) + (data[1,0]-data[3,0]) + (data[2,0]-data[3,0])

        return loss_first.cuda() + loss_second.cuda() + loss_third.cuda() + loss_fourth.cuda()

        #-----------------------之前版本-------------------------
        # if (data[0,1]-data[1,1]) !=0 \
        #     and ((data[0,1]-data[2,1]) + (data[1,1]-data[2,1])) !=0 \
        #     and ((data[0,1]-data[3,1])+(data[1,1]-data[3,1])+(data[2,1]-data[3,1]))!=0:
        #
        #     loss_first = torch.Tensor([0.])
        #
        #     loss_second = (data[0,1]-data[1,1]) / (data[0,1]-data[1,1]) * (data[0,0]-data[1,0])
        #
        #     loss_third = (data[0,1]-data[2,1]) / ((data[0,1]-data[2,1]) + (data[1,1]-data[2,1])) * (data[0,0]-data[2,0]) + \
        #                  (data[1,1]-data[2,1]) / ((data[0,1]-data[2,1]) + (data[1,1]-data[2,1])) * (data[1,0]-data[2,0])
        #
        #     loss_fourth = (data[0,1]-data[3,1]) / ((data[0,1]-data[3,1])+(data[1,1]-data[3,1])+(data[2,1]-data[3,1])) * (data[0,0]-data[3,0]) + \
        #                   (data[1,1]-data[3,1]) / ((data[0,1]-data[3,1])+(data[1,1]-data[3,1])+(data[2,1]-data[3,1])) * (data[1,0]-data[3,0]) + \
        #                   (data[2,1]-data[3,1]) / ((data[0,1]-data[3,1])+(data[1,1]-data[3,1])+(data[2,1]-data[3,1])) * (data[2,0]-data[3,0])
        #
        #     return loss_first.cuda() + loss_second.cuda() +  loss_third.cuda() + loss_fourth.cuda()
        # else:
        #     return torch.Tensor([0.]).cuda()


    # cls_label的shape为 【num_objects，】代表该框框所包含的真实类别值
    # container_embeddings的shape为【num_objects,256】
    def loss_long_tail_func(self, cls_label, container_embeddings, epoch, in_h, is_save):
        # cls_label中包含 0,1, 2, 3, 4共计5个类别，其分别对应 ['echinus', 'holothurian', 'starfish', 'scallop', 'waterweeds']

        if container_embeddings.shape[1] !=6:
            # 等号右侧：cls_label的shape为【num_objects，】，代表该框框所包含的真实类别值，并且这num_objects个数值仅包含0,1,2,3,4，分别代表['echinus', 'holothurian', 'starfish', 'scallop', 'waterweeds']
            # 等号左侧：利用 torch.where对cls_label操作，分别得到索引分别是0,1,2,3,4的行号，其shape为【num_echinus,】,其中num_echinus代表cls_label中等于0的个数，内容是行号，其余类似
            index_echinus = torch.where(cls_label==0)
            index_holothurian = torch.where(cls_label==1)
            index_starfish = torch.where(cls_label==2)
            index_scallop = torch.where(cls_label==3)

            # 等号右侧：利用 torch.where对cls_label操作，分别得到索引分别是0,1,2,3,4的行号，其shape为【num_echinus,】,其中num_echinus代表cls_label中等于0的个数，内容是行号，其余类似
            # 等号左侧：embedding_echinus的shape为【num_echinus， 256】，其余类似。
            embedding_echinus = container_embeddings[index_echinus].cpu()
            embedding_holothurian = container_embeddings[index_holothurian].cpu()
            embedding_starfish = container_embeddings[index_starfish].cpu()
            embedding_scallop = container_embeddings[index_scallop].cpu()


            if not is_save: # 如果不保存 方差var 和 角度 的话，则执行

                # 将每次进入循环中，每个类别的embedding与全局的global_embeddings进行cat操作
                # 等号右侧：embedding_echinus的shape为【num_echinus， 256】。其中num_echinus代表个数echinus的个数。其余变量类似。
                temp_embeddings_echinus = torch.cat([self.global_embeddings_echinus, embedding_echinus], dim=0)
                self.global_embeddings_echinus = temp_embeddings_echinus

                temp_embeddings_holothurian = torch.cat([self.global_embeddings_holothurian, embedding_holothurian], dim=0)
                self.global_embeddings_holothurian = temp_embeddings_holothurian

                temp_embeddings_starfish = torch.cat([self.global_embeddings_starfish, embedding_starfish], dim=0)
                self.global_embeddings_starfish = temp_embeddings_starfish

                temp_embeddings_scallop = torch.cat([self.global_embeddings_scallop, embedding_scallop], dim=0)
                self.global_embeddings_scallop = temp_embeddings_scallop

            else:

                path_for_angular_result = './logs/loss_file/' +'0-'+ str(in_h)+'x'+str(in_h)+'-Angular.xlsx'
                path_for_var_result = './logs/loss_file/'+'1-'+str(in_h)+'x'+str(in_h)+'-Var.txt'
                if  epoch == 0:
                    if os.path.exists(path_for_angular_result):
                        os.remove(path_for_angular_result)
                        print('删除 Angular.xlsx文件 成功',path_for_angular_result)
                    else:
                        print('没啥吊事')
                    if os.path.exists(path_for_var_result):
                        os.remove(path_for_var_result)
                        print('删除 -Var.txt文件 成功',path_for_var_result)
                    else:
                        print('不进行操作')

                    wb = Workbook()
                    wb.create_sheet(index=0,
                                    title='angle between center with items')  # 创建sheet1，其名称为 underwater_detection
                    wb.save(path_for_angular_result)
                    print('成功建立 Angular.xlsx文件', path_for_angular_result)

                else:
                    print('继续保留已存在 Angular.xlsx文件')

                print('准备向 Angular.xlsx 和 -Var.txt 写入内容')
                # 将损失函数保存在
                with open(path_for_var_result, 'a') as fff:
                    fff.write(str(self.var_echinus.data.numpy())+ ' '
                              +str(self.var_holothurian.data.numpy())+ ' '
                              +str(self.var_starfish.data.numpy())+ ' '
                              +str(self.var_scallop.data.numpy()))

                    fff.write('\n')
                    fff.close()

                wb = load_workbook(path_for_angular_result)  # 加载一个已存在的workbook对象
                wb1 = wb.active  # 激活默认的sheet表格，一般为第一个sheet

                wb1.cell(5 * epoch +  1, 1, str(epoch))  # 具体的含义分别是：行/列/数值
                wb1.cell(5 * epoch +  2, 1, str('echinus'))  # 具体的含义分别是：行/列/数值
                [wb1.cell(5 * epoch + 2, index + 2, str(item_angular.numpy())) for index, item_angular in enumerate(self.angular_between_items_with_center_echinus)]
                wb1.cell(5 * epoch +  3, 1, str('holothurian'))  # 具体的含义分别是：行/列/数值
                [wb1.cell(5 * epoch + 3, index + 2, str(item_angular.numpy())) for index, item_angular in enumerate(self.angular_between_items_with_center_holothurian)]
                wb1.cell(5 * epoch +  4, 1, str('starfish'))  # 具体的含义分别是：行/列/数值
                [wb1.cell(5 * epoch + 4, index + 2, str(item_angular.numpy())) for index, item_angular in enumerate(self.angular_between_items_with_center_starfish)]
                wb1.cell(5 * epoch +  5, 1, str('scallop'))  # 具体的含义分别是：行/列/数值
                [wb1.cell(5 * epoch + 5, index + 2, str(item_angular.numpy())) for index, item_angular in enumerate(self.angular_between_items_with_center_scallop)]

                wb.save(path_for_angular_result)  # 保存

                print('向 Angular.xlsx 和 -Var.txt 写入完毕，请指示！')

                # 在新的epoch中，将所收集到的每个类别embedding全部置位0
                self.global_embeddings_echinus = torch.zeros([1, self.num_in_features])
                self.global_embeddings_holothurian = torch.zeros([1, self.num_in_features])
                self.global_embeddings_starfish = torch.zeros([1, self.num_in_features])
                self.global_embeddings_scallop = torch.zeros([1, self.num_in_features])

                # 在新的epoch中，将所有 类别的中心 全部置位0
                self.cls_echinus_global_center = torch.zeros([1, self.num_in_features])
                self.cls_holothurian_global_center = torch.zeros([1, self.num_in_features])
                self.cls_starfish_global_center = torch.zeros([1, self.num_in_features])
                self.cls_scallop_global_center = torch.zeros([1, self.num_in_features])

                # 更新一下 ”不同类别“ 的每一个item与中心的夹角全部置位空集
                self.angular_between_items_with_center_echinus = []
                self.angular_between_items_with_center_holothurian = []
                self.angular_between_items_with_center_starfish = []
                self.angular_between_items_with_center_scallop = []

                return torch.Tensor([0.0]).cuda()

            # 计算 中心向量 与 每个类别中 样本向量 的 角度（删掉了其中为nan的样本点），angular_echinus的shape为【BBB，】，BBB的个数不知道
            # 有可能BBB的个数为1、2、3，需要说明的是，当BBB的个数为1的时候，不能计算方差，因此需要特别注意。
            angular_echinus, update_center_echinus = self.calculate_angular_between_each_item_and_center(self.global_embeddings_echinus, self.cls_echinus_global_center)
            angular_holothurian, update_center_holothurian = self.calculate_angular_between_each_item_and_center(self.global_embeddings_holothurian, self.cls_holothurian_global_center)
            angular_starfish, update_center_starfish = self.calculate_angular_between_each_item_and_center(self.global_embeddings_starfish, self.cls_starfish_global_center)
            angular_scallop, update_center_scallop = self.calculate_angular_between_each_item_and_center(self.global_embeddings_scallop, self.cls_scallop_global_center)


            # 更新一下 ”不同类别“ 的全局中心
            self.cls_echinus_global_center = update_center_echinus
            self.cls_holothurian_global_center = update_center_holothurian
            self.cls_starfish_global_center = update_center_starfish
            self.cls_scallop_global_center = update_center_scallop

            # 更新一下 ”不同类别“ 的每一个item与中心的夹角
            self.angular_between_items_with_center_echinus = angular_echinus
            self.angular_between_items_with_center_holothurian = angular_holothurian
            self.angular_between_items_with_center_starfish = angular_starfish
            self.angular_between_items_with_center_scallop = angular_scallop


            if angular_echinus.shape[0] >=2 and \
                angular_holothurian.shape[0] >=2 and \
                angular_starfish.shape[0] >=2 and \
                angular_scallop.shape[0] >=2:

                # 计算angular_echinus的方差，
                # 等号右侧：angular_echinus的shape为【BBB，】，BBB的个数不知道
                # 等号左侧：var_echinus的shape是【1，】
                self.var_echinus = torch.var(angular_echinus)
                self.var_holothurian = torch.var(angular_holothurian)
                self.var_starfish = torch.var(angular_starfish)
                self.var_scallop = torch.var(angular_scallop)


                # 等号右侧：var_echinus的shape是【1，】
                # 等号左侧：var_att的shape是【4,1】
                var_att = torch.Tensor([[self.var_echinus],
                                        [self.var_holothurian],
                                        [self.var_starfish],
                                        [self.var_scallop]])
                num_att = torch.Tensor([[angular_echinus.shape[0]],
                                        [angular_holothurian.shape[0]],
                                        [angular_starfish.shape[0]],
                                        [angular_scallop.shape[0]]])

                # 依次合并每一类方差属性（var_att） 及 每一类数量属性 （num_att）
                # 等号右侧：var_att的shape是 【4,1】 num_att的shape是【4,1】
                # 等号左侧：coll_att的shape是【4,2】
                coll_att = torch.cat([var_att, num_att], dim=1)

                # 等号左侧：var_att的shape是【4,1】，按照第一列的大小降序排序
                # 等号左侧：index_att的shape是【4，1】，表示的最大值在第几行，第二大值在第几行，第三大值在第几行 以及 第四大值在第几行
                index_att = torch.argsort(var_att, dim=0, descending=True)

                # 等号右侧：index_att的shape是【4，1】，表示的最大值在第几行，第二大值在第几行，第三大值在第几行 以及 第四大值在第几行
                # 等号左侧：decending_var_num按照index_att进行排序，最终decending_var_num的shape是【4,2】，形式如下：
                # 第一行：【最大的方差，类别数量】， 第二行：【第二大方差，类别数量】， 第三行：【第三大方差，类别数量】， 第四行：【第四大方差，类别数量】
                decending_var_num = coll_att[index_att].squeeze()

                # 计算长尾损失
                long_tail_loss = self.calculate_diff_cls_loss(decending_var_num)

                return long_tail_loss

            else:

                return torch.Tensor([0.0]).cuda()

        else:

            return torch.Tensor([0.0]).cuda()


    def cosine_loss_func(self, cosine_value, cls_label):
        # ------------------------陈廷凯自己设计类别损失方法---------------
        # 等号右侧： cls_label的shape为 【num_objects，】代表本尺度（13*13,26*26,52*52）下匹配上num_objects物体，同时每个值代表代表物体的真实类别
        # 等号左侧： index的shape为【num_objects，】，代表cls_label类别代号不等于-1的位置
        if cosine_value.shape[1]==5:

            # https://blog.csdn.net/GAIYA2050/article/details/104829140
            # 使用nn.CrossEntropyLoss前不需要经过softmax层
            # 等号右侧：cosine_value的shape为【num_objects, 5】代表cosine值，
            # 等号右侧：cls_label的shape为 【num_objects，】代表该框框所包含的真实类别值
            loss_cosine_sum = F.cross_entropy(cosine_value, cls_label.cuda(), reduction='sum')


            return loss_cosine_sum
        else:
            return torch.Tensor([0]).cuda()


    # 数据0: feature_map 的shape为 （bs,128, 13,13）表示检测头前面的一个特征层
    # 数据1：decode_output 的 shape为 (bs，3, 13, 13, 4) 其中4代表x,y,w,h,其是相对于特征图尺寸(预测值)。
    # 数据2：mask 的shape为(bs, 3, 13, 13) 表示存在目标的特征点
    # 数据3：tcls 的shape为 （bs, 3, 13, 13, 5），其中5代表5个类别，echinus, scallop, starfish, holothurian, waterweeds，其表示真实物体的类别
    def metric_cosine(self, feature_map, decode_output, mask, tcls):
        #print('tcls', type(tcls))
        # 等号右侧：decode_output的shape为(bs，3, 13, 13, 4)，其中4代表x,y,w,h,其是相对于特征图尺寸(预测值)。
        # 等号右侧：mask的shape为(bs, 3, 13, 13)
        # 等号左侧：exist_bboxs的shape（ZZZ,4），表示这batch size个图片中，存在ZZZ个真实物体。
        #exist_bboxs = decode_output[mask==1]

        # 定义两个容器来分别保存embeddings和clss
        container_embeddings = []
        container_clss = []

        # 循环遍历每一个batch size中的图片
        for b in range(feature_map.shape[0]):

            # 如果torch.sum大于1，则表示在第b张图片且在该尺度下存在真实的物体。因此应该获取第b张图片下，针对每一个真实物体的预测值。
            if torch.sum(mask[b])>=1:

                # 计算当前第b张图片下的框框
                # 等号右侧：decode_output的shape为(bs, 3, 13, 13, 4)，其中4代表x,y,w,h,其是相对于特征图尺寸(预测值)。
                # 等号右侧：mask         的shape为(bs, 3, 13, 13) 表示存在目标的特征点
                # 等号左侧：exist_bboxs  的shape为(DDD, 4),其中DDD表示在第b张图片下且在当前尺度（13*13,26*26，52*52）下真实物体的个数
                exist_bboxs = decode_output[b][mask[b] == 1]
                #print('exist_bboxs.shape', exist_bboxs.shape)

                # 等号右侧：tcls 的shape为（bs, 3, 13, 13, 5），其中5代表5个类别，echinus, scallop, starfish, holothurian, waterweeds，其表示真实物体的类别
                # 等号右侧：mask 的shape为 (bs, 3, 13, 13) 表示存在目标的特征点
                # 等号左侧：exist_cls 的shape为（DDD,5）,其中DDD表示在第b张图片下且在当前尺度（13*13,26*26，52*52）下真实物体的个数，5是一个独热编码，是哪个物体，则这个物体就为1,其余为0
                exist_cls = tcls[b][mask[b] == 1]
                #print('exist_cls.shape', exist_cls.shape)

                #container_embeddings_for_each_batch = []
                #container_clss_for_each_batch = []

                # 逐个框框进行统一生成128维度的向量
                # exist_bboxs  的shape为(DDD, 4),其中DDD表示在第b张图片下且在当前尺度（13*13,26*26，52*52）下真实物体的个数
                for index_box in range(exist_bboxs.shape[0]):

                    # 依次取出第index_box个预测框，注意到等号左侧输出exist_bbox的形式为x,y,w,h
                    exist_bbox = exist_bboxs[index_box]
                    # print('exist_bbox', exist_bbox.device)
                    # print(exist_bbox.shape)

                    # 等号右侧形式为：x,y,w,h的形式
                    # 等号左侧形式为：左上角（x1,y1）和右下角（x2,y2）
                    bbox_x1y1_x2y2 = [exist_bbox[0] - exist_bbox[2] / 2, exist_bbox[1] - exist_bbox[3] / 2,
                                      exist_bbox[0] + exist_bbox[2] / 2, exist_bbox[1] + exist_bbox[3] / 2]



                    # 获得 box信息
                    box_ROI = [max(bbox_x1y1_x2y2[0], 0),
                           max(bbox_x1y1_x2y2[1], 0),
                           min(bbox_x1y1_x2y2[2], feature_map.shape[2]-1),
                           min(bbox_x1y1_x2y2[3], feature_map.shape[3]-1)]

                    if box_ROI[0] > box_ROI[2] or box_ROI[1] > box_ROI[3]:
                        continue
                    else:

                        # 获得 ROI框所在的特征图
                        update_feature = feature_map[b].unsqueeze(dim=0).cpu()


                        # 更新 ROI的的左上角和右下角信息，将（x1,y1, x2,y2）转为torch.tensor类型
                        tensor_box = torch.tensor([[torch.tensor(coordinate).type_as(update_feature).item() for coordinate in box_ROI]])
                        #tensor_box = torch.as_tensor([[torch.tensor(coordinate).type_as(update_feature).item() for coordinate in box_ROI]])
                        #tensor_box = torch.as_tensor([[torch.Tensor(coordinate).item() for coordinate in box_ROI]], dtype=torch.float32)


                        # 将ROI所抠出来的区域划分为3×3的bin大小，对于每一个bin中再设置4个采样点提取特征，按照torchvision.ops.roi_align的思路，计算这四个的均值得到每一个bin的输出；
                        # 等号右侧：update_feature的形状为（1,256,13,13）或者（1,256,26,26）或者（1,256,52,52）
                        # 等号右侧：[update_box]的shape是（1,4）的两维tensor
                        # 等号左侧：pooled_regions的shape为（1,256,3,3）
                        # https://zhuanlan.zhihu.com/p/149859747
                        # 根据上面链接所提供的内容，output_size=(3, 3)代表要将所传入的[tensor_box]分成3*3=9份，9份中每一份通过4个点（sampling_ratio×sampling_ratio=4）得到；
                        pooled_regions = torchvision.ops.roi_align(update_feature, [tensor_box], output_size=(3, 3), sampling_ratio=1)

                        # 等号左侧：embedding的特征为（256,）
                        embedding = self.avg(pooled_regions).squeeze()

                        # 等号右侧 exist_cls 的shape为（DDD,5）,其中DDD表示在第b张图片下且在当前尺度（13*13,26*26，52*52）下真实物体的个数，5是一个独热编码，是哪个物体，则这个物体就为1,其余为0
                        # 等号左侧 利用torch.max函数会返回两个tensor，第一个tensor是每行的最大值；第二个tensor是每行最大值的索引
                        # https://www.jianshu.com/p/3ed11362b54f
                        cls_value, cls_indices = torch.max(exist_cls, 1)
                        cls_indices = cls_indices[index_box]

                        # 利用已经建好的容器来分别保存embedding和cls,
                        # 被收纳的embedding的特征shape为（128,）
                        container_embeddings.append(embedding)
                        # print('type(embedding)', type(embedding))

                        # 被收纳的cls的特征shape为（1，），其包含0,1,2,3,4这几个数字，代表每一类。
                        container_clss.append(cls_indices)


        # 最终遍历完所有的图片，可以获取shape为（num_objects，128）的embeddings和（num_objects，）的类别，在这种情况下，可以调用CosFace或者ArcFace了。
        # https://blog.csdn.net/weixin_40740309/article/details/114700259
        container_embeddings = torch.tensor([item.cpu().detach().numpy() for item in container_embeddings]).cuda()
        #print('len(container_embeddings)', type(container_embeddings), container_embeddings[0].type)

        #print('len(container_clss)', type(container_clss), container_clss[0].type)
        # 测试可以
        #cls_label = torch.from_numpy(np.array([item.cpu().detach().numpy() for item in container_clss]))
        # 测试可以
        # cls_label的shape为（num_objects，）
        cls_label = torch.from_numpy(np.array(container_clss))


        # # 防止出现本尺度（13*13,26*26,52*52，只能是其中的某一个尺度）下所有的batch_size张图片中都不包含真实物体, 所以container_embeddings.shape[0]必须大于0
        #print('陈廷凯想看一下在feature为%s下，有%s个 container_embeddings'%(feature_map.shape[2], container_embeddings.shape[0]))
        if container_embeddings.shape[0]>0:


        # F.normalize计算的是范数，默认是（p=2范数）(input, p=2, dim=1, eps=1e-12, out=None)
        # 等号右侧： F.normalize(input)的shape是【num_objects,128】 F.normalize(self.weight)的shape是【5, 128】，
        # 等号左侧： cosine的shape为【num_objects, 5】
        # 经过下面的公式，我们可以得到余弦角，而F.linear代表将两个输入F.normalize(input)和 F.normalize(self.weight)进行相乘，
            cosine = F.linear(F.normalize(container_embeddings), F.normalize(self.weight.cuda()))

            # 等号右侧：cosine的shape为【num_objects, 5】
            # 等号左侧：phi  的shape为【num_objects, 5】
            phi = cosine - self.m
            output = cosine * 1.0  # make backward works

            # num_objects代表所有batch_size张图片下且尺度为（13*13,26*26,52*52中的某一个）下，找到负责预测物体框框的个数
            num_objects = container_embeddings.shape[0]

            # 等号右侧：  phi的shape为 [num_objects, 5]； range(num_objects)代表从0,1，...num_objects; cls_label代表 num_objects 张图片属于的类别；所以等号右侧整体表示取出第XXX行，第XXX列的数据；
            # 然后赋值给output的相应位置
            # 等号右侧：  phi[range(batch_size), label]的shape为【num_objects，】
            # 等号左侧：  output的shape为【num_objects, 5】
            output[range(num_objects), cls_label.long()] = phi[range(num_objects), cls_label.long()]

            # 等号右侧：output的shape为【num_objects, 5】代表cosine值，
            # 等号右侧：cls_label的shape为 【num_objects，】代表该框框所包含的真实类别值
            #print('output.shape, cls_label.shape', output.shape, cls_label.shape)
            return output * self.s, cls_label, container_embeddings

        else:
            # 这是CTK随便定义了一个shape为【1,6】的张量，最后一维是6。主要的不同点在于上面return返回的是 【num_objects, 5】，最后一维度为5；
            # 也就是说当返回值为6维度的时候，表示所有batch size张图片中在本尺度（13*13,26*26,52*52）下，没有负责预测的先验框；
            return torch.Tensor([[0,1,2,3,4,5]]), torch.Tensor([[0,1,2,3,4,5]]), torch.Tensor([[0,1,2,3,4,5]])


    def get_target(self, target, anchors, in_w, in_h, ignore_threshold):
        # CTK  target的总数量为bs，对于bs的中每一个batch，其形式为N*5，其中N代表每张图片里面有多少个物体，5代表x,y,w,h和类别
        #      anchors 这里其实是scaled_anchors，是以416*416为衡量标准的anchor除以stride之后，获得shape是【9，2】
        #      in_w 特征图的宽度
        #      in_h 特征图的高度
        #-----------------------------------------------------#
        #   计算一共有多少张图片
        #-----------------------------------------------------#
        bs = len(target)

        #-------------------------------------------------------#
        #   获得当前特征层先验框所属的编号，方便后面对先验框筛选
        #-------------------------------------------------------#
        # CTK 确定 self.feature_length=[13、26、52], 因此借助于index函数，最终anchor_index的值只能为【0,1,2】、【3,4,5】和【6,7,8】任意一个方括号的三个值
        # 因此，若本次遍历的是13*13特征层，则anchor_index对应【0，1，2】
        #      若本次遍历的是26*26特征层，则anchor_index对应【3，4，5】
        #      若本次遍历的是52*52特征层，则anchor_index对应【6，7，8】
        anchor_index = [[0,1,2],[3,4,5],[6,7,8]][self.feature_length.index(in_w)]

        # CTK 确定 self.feature_length=[13、26、52], 因此借助于index函数，最终 subtract_index 的值只能为【0，3，6】中任意一个值
        # 因此，若本次遍历的是13*13特征层，则 subtract_index 对应【0】
        #      若本次遍历的是26*26特征层，则 subtract_index 对应【3】
        #      若本次遍历的是52*52特征层，则 subtract_index 对应【6】
        subtract_index = [0,3,6][self.feature_length.index(in_w)]
        #-------------------------------------------------------#
        #   创建全是0或者全是1的阵列
        #-------------------------------------------------------#
        mask = torch.zeros(bs, int(self.num_anchors/3), in_h, in_w, requires_grad=False)      # CTK self.num_anchors的长度为9，因此 mask 的shape为（bs, 3, 13,13）
        noobj_mask = torch.ones(bs, int(self.num_anchors/3), in_h, in_w, requires_grad=False) # CTK self.num_anchors的长度为9，因此 noobj_mask 的shape为（bs, 3, 13,13）

        tx = torch.zeros(bs, int(self.num_anchors/3), in_h, in_w, requires_grad=False)        # CTK self.num_anchors的长度为9，因此 tx 的shape为（bs, 3, 13,13）
        ty = torch.zeros(bs, int(self.num_anchors/3), in_h, in_w, requires_grad=False)        # CTK self.num_anchors的长度为9，因此 ty 的shape为（bs, 3, 13,13）
        tw = torch.zeros(bs, int(self.num_anchors/3), in_h, in_w, requires_grad=False)        # CTK self.num_anchors的长度为9，因此 tw 的shape为（bs, 3, 13,13）
        th = torch.zeros(bs, int(self.num_anchors/3), in_h, in_w, requires_grad=False)        # CTK self.num_anchors的长度为9，因此 th 的shape为（bs, 3, 13,13）
        tconf = torch.zeros(bs, int(self.num_anchors/3), in_h, in_w, requires_grad=False)     # CTK self.num_anchors的长度为9，因此 tconf 的shape为（bs, 3, 13,13）
        tcls = torch.zeros(bs, int(self.num_anchors/3), in_h, in_w, self.num_classes, requires_grad=False)  # CTK self.num_anchors的长度为9，因此 tcls 的shape为（bs, 3, 13,13, num_classes）

        box_loss_scale_x = torch.zeros(bs, int(self.num_anchors/3), in_h, in_w, requires_grad=False)   # CTK self.num_anchors的长度为9，因此 box_loss_scale_x 的shape为（bs, 3, 13,13）
        box_loss_scale_y = torch.zeros(bs, int(self.num_anchors/3), in_h, in_w, requires_grad=False)   # CTK self.num_anchors的长度为9，因此 box_loss_scale_y 的shape为（bs, 3, 13,13）

        # target的总数量为bs，对于bs的中每一个batch，其形式为N*5，其中N代表每一张图片里面有多少物体，5代表x,y,w,h和类别
        for b in range(bs):
            if len(target[b])==0:
                continue
            #-------------------------------------------------------#
            #   计算出正样本在特征层上的中心点
            #-------------------------------------------------------#
            # CTK 确定target的中心点 和 宽高 信息 在dataloader.py文件的YoloDataset 被转换成0~1的百分比，这里乘以特征图的宽高，获得以特征层为衡量标准的形式
            gxs = target[b][:, 0:1] * in_w  # gxs 的shape为【N,1】,其中N代表本图片中包含几个物体
            gys = target[b][:, 1:2] * in_h  # gys 的shape为【N,1】
            
            #-------------------------------------------------------#
            #   计算出正样本相对于特征层的宽高
            #-------------------------------------------------------#
            gws = target[b][:, 2:3] * in_w   # gws 的shape为【N,1】,其中N代表本图片中包含几个物体
            ghs = target[b][:, 3:4] * in_h   # ghs 的shape为【N,1】

            #-------------------------------------------------------#
            #   计算出正样本属于特征层的哪个特征点
            #-------------------------------------------------------#
            # CTK 确定 torch.floor表示向下取整的形式 torch.floor(100.72) :  100.0
            gis = torch.floor(gxs)
            gjs = torch.floor(gys)
            
            #-------------------------------------------------------#
            #   将真实框转换一个形式
            #   num_true_box, 4
            #-------------------------------------------------------#
            # gws                   shape为【N,1】
            # ghs                   shape为【N,1】
            # torch.zeros_like(gws) shape为【N,1】
            # torch.zeros_like(ghs) shape为【N,1】
            # CTK gt_box的形状【N, 4】,其中4代表[0,0,gws,ghs]
            gt_box = torch.FloatTensor(torch.cat([torch.zeros_like(gws), torch.zeros_like(ghs), gws, ghs], 1))

            #-------------------------------------------------------#
            #   将先验框转换一个形式
            #   9, 4
            #-------------------------------------------------------#
            # CTK 确定torch.zeros((self.num_anchors, 2))的shape是【9，2】
            # CKT 确定anchors                           的shape是【9，2】
            # CTK 确定 anchor_shapes                    的shape是【9，4】，其中 4 的形式为【0，0，scaled_anchors_w, scaled_anchors_h】
            anchor_shapes = torch.FloatTensor(torch.cat((torch.zeros((self.num_anchors, 2)), torch.FloatTensor(anchors)), 1))
            #-------------------------------------------------------#
            #   计算交并比
            #   num_true_box, 9
            #-------------------------------------------------------#
            # CTK 确定gt_box       的形状【N, 4】,其中 4 代表[0,0,gws,ghs]
            # CTK 确定anchor_shapes的形状【9，4】，其中 4 代表为【0，0，scaled_anchors_w, scaled_anchors_h】
            # CTK 确定 anch_ious   的形状【N，9】 其中N代表本图片中包含几个物体
            anch_ious = jaccard(gt_box, anchor_shapes)
            #print('anch_ious', anch_ious.shape)

            # 因此，若本次遍历的是13*13特征层，则anchor_index对应【0，1，2】
            #      若本次遍历的是26*26特征层，则anchor_index对应【3，4，5】
            #      若本次遍历的是52*52特征层，则anchor_index对应【6，7，8】
            # 等号右侧 anch_ious的shape为【N，9】，其中N代表本图片中包含几个物体
            # 等号左侧 anch_ious_select_three的shape为【N，3】
            #-------------------------------------------------------#
            #   计算重合度最大的先验框是哪个
            #   # CTK best_ns的shape是（N,）
            #-------------------------------------------------------#
            # best_ns的shape为【N，】，其中N代表本图片中包含几个物体
            best_ns = torch.argmax(anch_ious,dim=-1)

            for i, best_n in enumerate(best_ns):
                #  CTK  anchor_index 的值只能为【0,1,2】、【3,4,5】和【6,7,8】任意一个方括号的三个值
                #  因此，若本次遍历的是13*13特征层，则anchor_index对应【0，1，2】
                #       若本次遍历的是26*26特征层，则anchor_index对应【3，4，5】
                #       若本次遍历的是52*52特征层，则anchor_index对应【6，7，8】
                if best_n not in anchor_index:
                    #print(b, 'No match')
                    continue


                gi = gis[i].long()  # gis代表真值框的中心点x(这个x是相对于当前特征图的)，其shape为【N，1】。但是，gi的shape为【1,1】并且gi都是整数，
                gj = gjs[i].long()  # gjs代表真值框的中心点y(这个y是相对于当前特征图的)，其shape为【N，1】。但是，gj的shape为【1,1】并且gj都是整数，
                gx = gxs[i]         # gxs代表真值框的中心点x(这个x是相对于当前特征图的)，其shape为【N，1】。但是，gx的shape为【1,1】并且gx基本上均为小数。
                gy = gys[i]         # gys代表真值框的中心点x(这个x是相对于当前特征图的)，其shape为【N，1】。但是，gy的shape为【1,1】并且gy基本上均为小数。
                gw = gws[i]         # gws代表真值框的中心点x(这个w是相对于当前特征图的)，其shape为【N，1】。但是，gw的shape为【1,1】并且gw基本上均为小数。
                gh = ghs[i]         # ghs代表真值框的中心点x(这个h是相对于当前特征图的)，其shape为【N，1】。但是，gh的shape为【1,1】并且gh基本上均为小数。


                if (gj < in_h) and (gi < in_w): # 如果两者均满足，则代表真实框所在网格点在图像特征图以内
                    # 因此，若本次遍历的是13*13特征层，则 subtract_index 对应【0】,另外，等号右侧的 best_n是【0，1，2】中的一个
                    #      若本次遍历的是26*26特征层，则 subtract_index 对应【3】,另外，等号右侧的 best_n是【3，4，5】中的一个
                    #      若本次遍历的是52*52特征层，则 subtract_index 对应【6】,另外，等号右侧的 best_n是【6，7，8】中的一个
                    best_n = best_n - subtract_index

                    #----------------------------------------#
                    #   noobj_mask代表无目标的特征点
                    #----------------------------------------#
                    # 注意到 noobj_mask的定义为： torch.ones(bs, int(self.num_anchors/3), in_h, in_w, requires_grad=False)
                    # 所以，需要将gj（代表y，也就是高度值h）放在前面， 同时将gi（代表x，也就是宽度值w）放在后面
                    noobj_mask[b, best_n, gj, gi] = 0        # noobj_mask 的shape为（bs, 3, 13,13），初次定义的时候为全为 1
                    #----------------------------------------#
                    #   mask代表有目标的特征点
                    #----------------------------------------#
                    mask[b, best_n, gj, gi] = 1              # mask 的shape为（bs, 3, 13,13），初次定义全为 0
                    #----------------------------------------#
                    #   tx、ty代表中心调整参数的真实值
                    #----------------------------------------#
                    tx[b, best_n, gj, gi] = gx - gi.float()  # tx 的shape为（bs, 3, 13,13）
                    ty[b,               best_n, gj, gi] = gy - gj.float()  # ty 的shape为（bs, 3, 13,13）

                    #----------------------------------------#
                    #   tw、th代表宽高调整参数的真实值
                    #----------------------------------------#
                    # 等号右侧：anchors 这里其实是scaled_anchors，是以416 * 416为衡量标准的anchor除以stride之后，获得shape是【9，2】（w, h）
                    #  若  本次遍历的是13*13特征层，则 subtract_index 对应【0】,另外，等号右侧的 best_n是【0，1，2】中的一个
                    #      若本次遍历的是26*26特征层，则 subtract_index 对应【3】,另外，等号右侧的 best_n是【3，4，5】中的一个
                    #      若本次遍历的是52*52特征层，则 subtract_index 对应【6】,另外，等号右侧的 best_n是【6，7，8】中的一个
                    tw[b, best_n, gj, gi] = math.log(gw / anchors[best_n + subtract_index][0])  # tw 的shape为（bs, 3, 13,13）
                    th[b, best_n, gj, gi] = math.log(gh / anchors[best_n + subtract_index][1])  # th 的shape为（bs, 3, 13,13）
                    #----------------------------------------#
                    #   用于获得xywh的比例
                    #   大目标loss权重小，小目标loss权重大
                    #----------------------------------------#
                    # target的总数量为bs，对于bs的中每一个batch，其形式为N * 5，其中N代表每张图片里面有多少个物体，5代表x, y, w, h和类别
                    # 这里获取的target[b][i, 2]和target[b][i, 3]都是已经被归一化到【0,1】之间，
                    # 引用，train_TK_edit里一句话 targets是一个batch的目标位置信息（每张图片中所包含物体 左上角 和 右下角 坐标信息依照原始图片大小（随便举个例子，例如，325*856，总之，这里并不是416*416）被归一化0-1之间
                    box_loss_scale_x[b, best_n, gj, gi] = target[b][i, 2]  # box_loss_scale_x 的shape为（bs, 3, 13,13）
                    box_loss_scale_y[b, best_n, gj, gi] = target[b][i, 3]  # box_loss_scale_y 的shape为（bs, 3, 13,13）
                    #----------------------------------------#
                    #   tconf代表物体置信度
                    #----------------------------------------#
                    tconf[b, best_n, gj, gi] = 1             # tconf 的shape为（bs, 3, 13,13）
                    #----------------------------------------#
                    #   tcls代表种类置信度
                    #target的总数量为bs，对于bs的中每一个batch，其形式为N * 5，其中N代表每张图片里面有多少个物体，5代表x, y, w, h和类别
                    #----------------------------------------#
                    # tcls = torch.zeros(bs, int(self.num_anchors/3), in_h, in_w, self.num_classes, requires_grad=False)其中，# CTK self.num_anchors的长度为9，
                    # 因此 tcls 的shape为torch.zeros（bs, 3, 13,13, num_classes）
                    tcls[b, best_n, gj, gi, int(target[b][i, 4])] = 1
                else:
                    print('Step {0} out of bound'.format(b))
                    print('gj: {0}, height: {1} | gi: {2}, width: {3}'.format(gj, in_h, gi, in_w))
                    continue

        return mask, noobj_mask, tx, ty, tw, th, tconf, tcls, box_loss_scale_x, box_loss_scale_y

    def get_ignore(self,prediction,target,scaled_anchors,in_w, in_h,noobj_mask):

        #---------------------------------------------------------------#
        #   将预测结果进行解码，判断预测结果和真实值的重合程度
        #   如果重合程度过大则忽略，因为这些特征点属于预测比较准确的特征点
        #   作为负样本不合适
        #----------------------------------------------------------------#
        # CTK  prediction的形式为 （b, 3, 13,13, (1+ 4+ num_class)）
        # CTK  target的总数量为bs，对于bs的中每一个batch，其形式为【N，5】，其中N代表一张图片里面包含物体的个数，5代表x,y,w,h和类别，另外 targets是一个batch的目标
        #       位置信息（每张图片中所包含物体 左上角 和 右下角 坐标信息依照原始图片大小（随便举个例子，例如，325*856，总之，这里并不是416*416）被归一化0-1之间
        # CTK  scaled_anchors的shape为【9，2】，scaled_anchors大小是相对于特征层的
        # in_w 特征层的宽
        # in_h 特征层的高
        # 等号右侧的noobj_mask 的shape为（bs, 3, 13,13）


        #-----------------------------------------------------#
        #   计算一共有多少张图片
        #-----------------------------------------------------#
        bs = len(target)
        #-------------------------------------------------------#
        #   获得当前特征层先验框所属的编号，方便后面对先验框筛选
        #-------------------------------------------------------#
        # CTK  确定 self.feature_length=[13、26、52], 因此借助于index函数，最终anchor_index的值只能为【0,1,2】、【3,4,5】和【6,7,8】任意一个方括号的三个值
        # 因此，若本次遍历的是13*13特征层，则anchor_index对应【0，1，2】
        #      若本次遍历的是26*26特征层，则anchor_index对应【3，4，5】
        #      若本次遍历的是52*52特征层，则anchor_index对应【6，7，8】
        anchor_index = [[0,1,2],[3,4,5],[6,7,8]][self.feature_length.index(in_w)]

        # 等号右侧 scaled_anchors的shape为【9，2】，scaled_anchors大小是相对于特征层的
        # CTK 输入：scaled_anchors的shape是【9，2】
        #     输出：scaled_anchors的shape是【3，2】
        scaled_anchors = np.array(scaled_anchors)[anchor_index]

        # prediction的shape为（b, 3, 13,13, (1+ 4+ num_class)）
        # 先验框的中心位置的调整参数  # x和y的shape均为[bs, 3, 13, 13]
        x = torch.sigmoid(prediction[..., 0])  
        y = torch.sigmoid(prediction[..., 1])

        # 先验框的宽高调整参数       # w和h的shape均为[bs, 3, 13, 13]
        w = prediction[..., 2]  # Width
        h = prediction[..., 3]  # Height

        FloatTensor = torch.cuda.FloatTensor if x.is_cuda else torch.FloatTensor
        LongTensor = torch.cuda.LongTensor if x.is_cuda else torch.LongTensor

        # 生成网格，先验框中心，网格左上角
        # CTK   grid_x 的形状为（bs,3,13,13）
        grid_x = torch.linspace(0, in_w - 1, in_w).repeat(in_h, 1).repeat(
            int(bs*self.num_anchors/3), 1, 1).view(x.shape).type(FloatTensor)
        # CTK  grid_y 的形状为（bs,3,13,13）
        grid_y = torch.linspace(0, in_h - 1, in_h).repeat(in_w, 1).t().repeat(
            int(bs*self.num_anchors/3), 1, 1).view(y.shape).type(FloatTensor)

        # 生成先验框的宽高
        anchor_w = FloatTensor(scaled_anchors).index_select(1, LongTensor([0]))
        anchor_h = FloatTensor(scaled_anchors).index_select(1, LongTensor([1]))

        # CTK anchor_w 的shape为（bs,3,13,13）
        anchor_w = anchor_w.repeat(bs, 1).repeat(1, 1, in_h * in_w).view(w.shape)
        # CTK anchor_h 的shape为（bs,3,13,13）
        anchor_h = anchor_h.repeat(bs, 1).repeat(1, 1, in_h * in_w).view(h.shape)
        
        #-------------------------------------------------------#
        #   计算调整后的先验框中心与宽高
        #-------------------------------------------------------#
        # pred_boxes的shape为（b, 3, 13, 13, 4)
        pred_boxes = FloatTensor(prediction[..., :4].shape)

        # # 先验框的中心位置的调整参数  # x和y的shape均为[bs, 3, 13, 13]
        pred_boxes[..., 0] = x.data + grid_x
        pred_boxes[..., 1] = y.data + grid_y

        # # 先验框的宽高调整参数       # w和h的shape均为[bs, 3, 13, 13]
        pred_boxes[..., 2] = torch.exp(w.data) * anchor_w
        pred_boxes[..., 3] = torch.exp(h.data) * anchor_h

        for b in range(bs):
            # pred_boxes的形式为（bs, 3, 13, 13, 4)，经过pred_boxes【i】索引之后变成（3, 13, 13, 4）
            pred_boxes_for_ignore = pred_boxes[b]

            # CTK 将pred_boxes_for_ignore（1,3,13,13,4）转换为（1*3*13*13，4）
            pred_boxes_for_ignore = pred_boxes_for_ignore.view(-1, 4)
            #-------------------------------------------------------#
            #   计算真实框，并把真实框转换成相对于特征层的大小
            #   gt_box      num_true_box, 4
            #-------------------------------------------------------#
            # CTK  target的总数量为bs，对于bs的中每一个batch，其形式为【N，5】，其中N代表一张图片中包含物体的个数，5代表x,y,w,h和类别
            if len(target[b]) > 0:
                gx = target[b][:, 0:1] * in_w
                gy = target[b][:, 1:2] * in_h
                gw = target[b][:, 2:3] * in_w
                gh = target[b][:, 3:4] * in_h

                # CTK gx、gy、gw、gh的形式都为（N,1）,经过torch.cat操作，gt_box变为（N,4）
                gt_box = torch.FloatTensor(torch.cat([gx, gy, gw, gh],-1)).type(FloatTensor)

                #-------------------------------------------------------#
                #   计算交并比
                #   CTK anch_ious 的shape为（num_true_box, num_anchors）
                #-------------------------------------------------------#
                # CTK  确定gt_box的形状为（N,4）, pred_boxes_for_ignore的形状为（1*3*13*13，4）,anch_ious的形式为（N, 1*3*13*13）
                anch_ious = jaccard(gt_box, pred_boxes_for_ignore)

                #-------------------------------------------------------#
                #   每个先验框对应真实框的最大重合度
                #   anch_ious_max   num_anchors
                #-------------------------------------------------------#
                # CTK torch.max 返回每一列中最大值的那个元素，且返回索引（返回最大元素在这一列的行索引）,anch_ious_max的形式为（1*3*13*13，）
                anch_ious_max, _ = torch.max(anch_ious,dim=0)

                # CTK 确定输入anch_ious_max的形式为（1*3*13*13，）， 并且 pred_boxes[i].size()[:3]索引之后变成（3, 13, 13）
                # CTK 确定输出anch_ious_max的形式为（3, 13, 13）
                anch_ious_max = anch_ious_max.view( pred_boxes[b].size()[:3] )


                # CTK noobj_mask的shape为（bs, 3, 13,13），经过noobj_mask[i]变为（3, 13,13）
                # 本行代码的意思是将大于self.ignore_threshold（0.5）的设置为 0，（也就是说 先验框 与其中任意一个 真值框的IOU大于0.5，代表这个先验框有预测物体的能力。）
                # 注意： noobj_mask这个变量的初始值全为 1
                noobj_mask[b][anch_ious_max > self.ignore_threshold] = 0
        return noobj_mask


def rand(a=0, b=1):
    return np.random.rand()*(b-a) + a


class Generator(object):
    def __init__(self,batch_size,
                 train_lines, image_size,
                 ):
        
        self.batch_size = batch_size
        self.train_lines = train_lines
        self.train_batches = len(train_lines)
        self.image_size = image_size
        
    def get_random_data(self, annotation_line, input_shape, jitter=.3, hue=.1, sat=1.5, val=1.5, random=True):
        '''r实时数据增强的随机预处理'''
        line = annotation_line.split()
        image = Image.open(line[0])
        iw, ih = image.size
        h, w = input_shape
        box = np.array([np.array(list(map(int,box.split(',')))) for box in line[1:]])

        if not random:
            scale = min(w/iw, h/ih)
            nw = int(iw*scale)
            nh = int(ih*scale)
            dx = (w-nw)//2
            dy = (h-nh)//2

            image = image.resize((nw,nh), Image.BICUBIC)
            new_image = Image.new('RGB', (w,h), (128,128,128))
            new_image.paste(image, (dx, dy))
            image_data = np.array(new_image, np.float32)

            # 调整目标框坐标
            box_data = np.zeros((len(box), 5))
            if len(box) > 0:
                np.random.shuffle(box)
                box[:, [0, 2]] = box[:, [0, 2]] * nw / iw + dx
                box[:, [1, 3]] = box[:, [1, 3]] * nh / ih + dy
                box[:, 0:2][box[:, 0:2] < 0] = 0
                box[:, 2][box[:, 2] > w] = w
                box[:, 3][box[:, 3] > h] = h
                box_w = box[:, 2] - box[:, 0]
                box_h = box[:, 3] - box[:, 1]
                box = box[np.logical_and(box_w > 1, box_h > 1)]  # 保留有效框
                box_data = np.zeros((len(box), 5))
                box_data[:len(box)] = box

            return image_data, box_data

        # resize image
        new_ar = w/h * rand(1-jitter,1+jitter)/rand(1-jitter,1+jitter)
        scale = rand(.25, 2)
        if new_ar < 1:
            nh = int(scale*h)
            nw = int(nh*new_ar)
        else:
            nw = int(scale*w)
            nh = int(nw/new_ar)
        image = image.resize((nw,nh), Image.BICUBIC)

        # place image
        dx = int(rand(0, w-nw))
        dy = int(rand(0, h-nh))
        new_image = Image.new('RGB', (w,h), (128,128,128))
        new_image.paste(image, (dx, dy))
        image = new_image

        # flip image or not
        flip = rand()<.5
        if flip: image = image.transpose(Image.FLIP_LEFT_RIGHT)

        # distort image
        hue = rand(-hue, hue)
        sat = rand(1, sat) if rand()<.5 else 1/rand(1, sat)
        val = rand(1, val) if rand()<.5 else 1/rand(1, val)
        x = cv2.cvtColor(np.array(image,np.float32)/255, cv2.COLOR_RGB2HSV)
        x[..., 0] += hue*360
        x[..., 0][x[..., 0]>1] -= 1
        x[..., 0][x[..., 0]<0] += 1
        x[..., 1] *= sat
        x[..., 2] *= val
        x[x[:,:, 0]>360, 0] = 360
        x[:, :, 1:][x[:, :, 1:]>1] = 1
        x[x<0] = 0
        image_data = cv2.cvtColor(x, cv2.COLOR_HSV2RGB)*255

        # correct boxes
        box_data = np.zeros((len(box),5))
        if len(box)>0:
            np.random.shuffle(box)
            box[:, [0,2]] = box[:, [0,2]]*nw/iw + dx
            box[:, [1,3]] = box[:, [1,3]]*nh/ih + dy
            if flip: box[:, [0,2]] = w - box[:, [2,0]]
            box[:, 0:2][box[:, 0:2]<0] = 0
            box[:, 2][box[:, 2]>w] = w
            box[:, 3][box[:, 3]>h] = h
            box_w = box[:, 2] - box[:, 0]
            box_h = box[:, 3] - box[:, 1]
            box = box[np.logical_and(box_w>1, box_h>1)] # discard invalid box
            box_data = np.zeros((len(box),5))
            box_data[:len(box)] = box

        return image_data, box_data

    def generate(self, train=True):
        while True:
            shuffle(self.train_lines)
            lines = self.train_lines
            inputs = []
            targets = []
            for annotation_line in lines:  
                if train:
                    img,y=self.get_random_data(annotation_line, self.image_size[0:2])
                else:
                    img,y=self.get_random_data(annotation_line, self.image_size[0:2], random=False)

                if len(y)!=0:
                    boxes = np.array(y[:,:4],dtype=np.float32)
                    boxes[:,0] = boxes[:,0]/self.image_size[1]
                    boxes[:,1] = boxes[:,1]/self.image_size[0]
                    boxes[:,2] = boxes[:,2]/self.image_size[1]
                    boxes[:,3] = boxes[:,3]/self.image_size[0]

                    boxes = np.maximum(np.minimum(boxes,1),0)
                    boxes[:,2] = boxes[:,2] - boxes[:,0]
                    boxes[:,3] = boxes[:,3] - boxes[:,1]
    
                    boxes[:,0] = boxes[:,0] + boxes[:,2]/2
                    boxes[:,1] = boxes[:,1] + boxes[:,3]/2
                    y = np.concatenate([boxes,y[:,-1:]],axis=-1)
                    
                img = np.array(img,dtype = np.float32)

                inputs.append(np.transpose(img/255.0,(2,0,1)))                  
                targets.append(np.array(y,dtype = np.float32))
                if len(targets) == self.batch_size:
                    tmp_inp = np.array(inputs)
                    tmp_targets = targets
                    inputs = []
                    targets = []
                    yield tmp_inp, tmp_targets


